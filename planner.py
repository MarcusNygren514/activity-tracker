"""
Activity Tracker – Resursplaneringsläsare
Läser Oaks Resursplanering.xlsm och returnerar planerade aktiviteter per vecka.
Kopierar filen till temp för att undvika låsningsproblem.
Skriver ALDRIG till originalfilen.
"""

import re
import shutil
import logging
import tempfile
from contextlib import contextmanager
from datetime import datetime, timedelta
from pathlib import Path

log = logging.getLogger("planner")

_WEEK_RE   = re.compile(r'^V(\d{2})(\d{2})$')
_PS_CODE_RE = re.compile(r'^([PS]\d{5})')


def _week_to_dates(year_short: int, week: int):
    year   = 2000 + year_short
    monday = datetime.strptime(f"{year}-W{week:02d}-1", "%Y-W%W-%w")
    return monday, monday + timedelta(days=6)


def _parse_hours(val) -> int:
    if val is None:
        return 0
    try:
        return round(float(val))
    except (ValueError, TypeError):
        return 0


@contextmanager
def _open_worksheet(file_path: str):
    """Kopierar filen till temp, öppnar med openpyxl och ger tillbaka aktiva bladet."""
    src = Path(file_path)
    if not src.exists():
        raise FileNotFoundError(f"Filen hittades inte: {file_path}")
    tmp = Path(tempfile.mktemp(suffix=src.suffix))
    shutil.copy2(src, tmp)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(tmp), read_only=True, data_only=True, keep_vba=False)
        try:
            yield wb.active
        finally:
            wb.close()
    finally:
        try:
            tmp.unlink()
        except Exception:
            pass


def read_all_projects(file_path: str) -> dict:
    """
    Läser hela planeringsfilen och returnerar alla projekt som {kod: namn},
    oavsett resurs eller datum.
    """
    with _open_worksheet(file_path) as ws:
        headers = {i: (str(c.value).strip() if c.value else "") for i, c in enumerate(ws[1])}
        col_project = next((i for i, v in headers.items() if v == "PROJEKT"), 1)

        registry = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if col_project >= len(row):
                continue
            val = str(row[col_project]).strip() if row[col_project] else ""
            if not val or val in ("None", "PROJEKT"):
                continue
            m = _PS_CODE_RE.match(val)
            if m:
                code = m.group(1)
                registry[code] = val[len(code):].strip(" -–")
        return registry


def read_planning(file_path: str, resource_name: str, from_date: datetime, to_date: datetime) -> list[dict]:
    """
    Läser planeringsfilen och returnerar aktiviteter för resource_name
    inom [from_date, to_date].

    Returnerar lista med:
    {
        "project":    str,
        "activity":   str,
        "week":       str,   # "V2619"
        "week_start": str,   # ISO-datum för veckans måndag
        "week_end":   str,   # ISO-datum för veckans söndag
        "hours":      float,
    }
    """
    with _open_worksheet(file_path) as ws:
        headers   = {}
        week_cols = {}

        for i, cell in enumerate(ws[1]):
            val = str(cell.value).strip() if cell.value else ""
            headers[i] = val
            m = _WEEK_RE.match(val)
            if m:
                week_cols[val] = i

        col_project  = next((i for i, v in headers.items() if v == "PROJEKT"),  1)
        col_activity = next((i for i, v in headers.items() if v == "AKTIVITET"), 2)
        col_resource = next((i for i, v in headers.items() if v == "RESURS"),    5)

        relevant_weeks = {}
        for week_name, col_idx in week_cols.items():
            m = _WEEK_RE.match(week_name)
            if not m:
                continue
            monday, sunday = _week_to_dates(int(m.group(1)), int(m.group(2)))
            if monday <= to_date and sunday >= from_date:
                relevant_weeks[week_name] = (col_idx, monday, sunday)

        if not relevant_weeks:
            return []

        results = []
        resource_lower = resource_name.strip().lower()

        for row in ws.iter_rows(min_row=2, values_only=True):
            resource_val = str(row[col_resource]).strip() if row[col_resource] else ""
            if resource_val.lower() != resource_lower:
                continue

            project  = str(row[col_project]).strip()  if row[col_project]  else ""
            activity = str(row[col_activity]).strip() if row[col_activity] else ""

            if not project or project in ("None", "PROJEKT"):
                continue

            for week_name, (col_idx, monday, sunday) in relevant_weeks.items():
                hours = _parse_hours(row[col_idx + 1]) if col_idx + 1 < len(row) else 0.0
                if hours <= 0:
                    continue
                results.append({
                    "project":    project,
                    "activity":   activity,
                    "week":       week_name,
                    "week_start": monday.strftime("%Y-%m-%d"),
                    "week_end":   sunday.strftime("%Y-%m-%d"),
                    "hours":      hours,
                })

        return results
